import openpyxl
import datetime


class ExcelOrder(object):
    def __init__(self, filepath):
        self.filepath = filepath
        try:
            self.wb = openpyxl.load_workbook(filename=self.filepath)
        except:
            print("Ошибка загрузки Excel файла")
        self.ws = self.wb.active
        self.number = ''
        self.date_time = ''
        self.shop = ''
        self.client = ''
        self.bonus_card = ''
        self.price = ''
        self.worker = ''

    def set_order(self, string_number):
        self.number = self.ws['A'+str(string_number)].value
        self.date_time = self.ws['B'+str(string_number)].value
        self.date_time = datetime.datetime.strptime(self.date_time, '%d.%m.%Y %H:%M:%S')
        self.shop = self.ws['C'+str(string_number)].value
        self.client = self.ws['D'+str(string_number)].value
        self.bonus_card = self.ws['E'+str(string_number)].value
        self.price = self.ws['F'+str(string_number)].value
        self.worker = self.ws['G'+str(string_number)].value
        return self

# ex = ExcelOrder()
# for i in range(2, 1000, 1):
#     order = ex.set_order(i)